#!/usr/bin/env python
"""Measure static angle of repose and spreading distance from an EDEM-like screenshot."""

from __future__ import annotations

import argparse
import json
import math
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Iterable

from PIL import Image, ImageDraw, ImageFont


@dataclass
class CropSpec:
    side: str
    filename: str
    left: int
    top: int
    right: int
    bottom: int

    @property
    def width(self) -> int:
        return self.right - self.left

    @property
    def height(self) -> int:
        return self.bottom - self.top


@dataclass
class FitResult:
    source_image: str
    binary_image: str
    side: str
    width_px: int
    height_px: int
    threshold: int
    boundary_points: int
    slope_tan_repose_angle: float
    abs_slope_tan_repose_angle: float
    intercept_y_up: float
    repose_angle_deg: float
    r_squared: float


def is_black(rgb: tuple[int, int, int], threshold: int = 35) -> bool:
    r, g, b = rgb
    return r < threshold and g < threshold and b < threshold


def is_cylinder_gray(rgb: tuple[int, int, int]) -> bool:
    r, g, b = rgb
    return 120 < r < 245 and 120 < g < 245 and 120 < b < 245 and abs(r - g) < 8 and abs(g - b) < 8


def contiguous_runs(values: Iterable[int]) -> list[tuple[int, int]]:
    runs: list[tuple[int, int]] = []
    start = prev = None
    for value in values:
        if start is None:
            start = prev = value
        elif value == prev + 1:
            prev = value
        else:
            runs.append((start, prev))
            start = prev = value
    if start is not None:
        runs.append((start, prev))
    return runs


def detect_cylinder(img: Image.Image) -> tuple[int, int, int]:
    width, height = img.size
    y_limit = min(450, height)
    gray_counts: list[tuple[int, int]] = []
    px = img.load()
    for x in range(width):
        count = 0
        for y in range(y_limit):
            if is_cylinder_gray(px[x, y]):
                count += 1
        gray_counts.append((x, count))
    center_min = width * 0.25
    center_max = width * 0.75
    gray_cols = [x for x, count in gray_counts if count > 100]
    runs = [run for run in contiguous_runs(gray_cols) if run[1] > run[0]]
    cylinder_runs = [
        run
        for run in runs
        if run[1] - run[0] + 1 >= 5 and center_min <= (run[0] + run[1]) / 2 <= center_max
    ]
    if not cylinder_runs:
        center_counts = [(x, count) for x, count in gray_counts if center_min <= x <= center_max]
        max_center_count = max((count for _x, count in center_counts), default=0)
        adaptive_threshold = max(10, int(max_center_count * 0.6))
        adaptive_cols = [x for x, count in center_counts if count >= adaptive_threshold]
        cylinder_runs = [
            run
            for run in contiguous_runs(adaptive_cols)
            if run[1] - run[0] + 1 >= 5 and center_min <= (run[0] + run[1]) / 2 <= center_max
        ]
    if not cylinder_runs and not runs:
        raise ValueError("Could not detect semi-transparent cylinder.")
    if cylinder_runs:
        cyl_left = min(run[0] for run in cylinder_runs)
        cyl_right = max(run[1] for run in cylinder_runs)
    else:
        cyl_left, cyl_right = max(runs, key=lambda run: run[1] - run[0])

    gray_points_y = []
    for x in range(cyl_left, cyl_right + 1):
        for y in range(min(500, height)):
            if is_cylinder_gray(px[x, y]):
                gray_points_y.append(y)
    if not gray_points_y:
        raise ValueError("Could not detect cylinder vertical extent.")
    cyl_bottom = max(gray_points_y)
    return cyl_left, cyl_right, cyl_bottom


def detect_soil_bbox(img: Image.Image) -> tuple[int, int, int, int]:
    width, height = img.size
    px = img.load()
    x_min, x_max = max(0, width // 8), min(width, width * 7 // 8)
    y_min, y_max = height * 45 // 100, min(height, height * 85 // 100)
    points: list[tuple[int, int]] = []
    for y in range(y_min, y_max):
        for x in range(x_min, x_max):
            if is_black(px[x, y]):
                points.append((x, y))
    if not points:
        raise ValueError("Could not detect black soil pile.")
    return (
        min(x for x, _ in points),
        min(y for _, y in points),
        max(x for x, _ in points),
        max(y for _, y in points),
    )


def max_black_y_at_column(img: Image.Image, x: int, y_min: int, y_max: int) -> int:
    px = img.load()
    ys = [y for y in range(y_min, y_max + 1) if is_black(px[x, y])]
    if not ys:
        raise ValueError(f"No black pixels found at x={x}.")
    return max(ys)


def make_crops(img: Image.Image, prefix: str, local_dir: Path) -> tuple[list[CropSpec], tuple[int, int, int], tuple[int, int, int, int]]:
    cyl_left, cyl_right, cyl_bottom = detect_cylinder(img)
    soil_left, soil_top, soil_right, soil_bottom = detect_soil_bbox(img)

    left_bottom = max_black_y_at_column(img, soil_left, soil_top, soil_bottom) + 1
    right_bottom = max_black_y_at_column(img, soil_right, soil_top, soil_bottom) + 1

    crops = [
        CropSpec("left", f"{prefix}1.png", soil_left, cyl_bottom, cyl_left, left_bottom),
        CropSpec("right", f"{prefix}2.png", cyl_right, cyl_bottom, soil_right + 1, right_bottom),
    ]
    for crop in crops:
        img.crop((crop.left, crop.top, crop.right, crop.bottom)).save(local_dir / crop.filename)
    return crops, (cyl_left, cyl_right, cyl_bottom), (soil_left, soil_top, soil_right, soil_bottom)


def binarize_and_fit(source: Path, output: Path, threshold: int, side: str) -> FitResult:
    gray = Image.open(source).convert("L")
    width, height = gray.size
    binary = Image.new("L", (width, height), 255)
    gp = gray.load()
    bp = binary.load()
    for y in range(height):
        for x in range(width):
            bp[x, y] = 0 if gp[x, y] < threshold else 255
    binary.save(output)

    boundary: list[tuple[float, float]] = []
    for x in range(width):
        for y in range(height):
            if bp[x, y] == 0:
                boundary.append((float(x), float(y)))
                break
    if len(boundary) < 2:
        raise ValueError(f"Not enough boundary points in {source.name}.")

    xs = [x for x, _ in boundary]
    ys_up = [float(height - 1) - y for _, y in boundary]
    n = len(xs)
    mean_x = sum(xs) / n
    mean_y = sum(ys_up) / n
    denom = sum((x - mean_x) ** 2 for x in xs)
    slope = sum((x - mean_x) * (y - mean_y) for x, y in zip(xs, ys_up)) / denom
    intercept = mean_y - slope * mean_x
    fitted = [slope * x + intercept for x in xs]
    ss_res = sum((y - yh) ** 2 for y, yh in zip(ys_up, fitted))
    ss_tot = sum((y - mean_y) ** 2 for y in ys_up)
    r2 = 1 - ss_res / ss_tot if ss_tot else 1.0
    angle = math.degrees(math.atan(abs(slope)))

    return FitResult(
        source_image=source.name,
        binary_image=output.name,
        side=side,
        width_px=width,
        height_px=height,
        threshold=threshold,
        boundary_points=n,
        slope_tan_repose_angle=slope,
        abs_slope_tan_repose_angle=abs(slope),
        intercept_y_up=intercept,
        repose_angle_deg=angle,
        r_squared=r2,
    )


def load_font(size: int) -> ImageFont.ImageFont:
    for fp in [r"C:\Windows\Fonts\msyh.ttc", r"C:\Windows\Fonts\simhei.ttf", r"C:\Windows\Fonts\simsun.ttc"]:
        try:
            return ImageFont.truetype(fp, size)
        except Exception:
            pass
    return ImageFont.load_default()


def draw_dashed_line(draw: ImageDraw.ImageDraw, p1: tuple[float, float], p2: tuple[float, float], width: int = 4) -> None:
    red = (220, 0, 0)
    dash, gap = 22, 12
    x1, y1 = p1
    x2, y2 = p2
    dx, dy = x2 - x1, y2 - y1
    length = (dx * dx + dy * dy) ** 0.5
    if length == 0:
        return
    ux, uy = dx / length, dy / length
    dist = 0
    while dist < length:
        end = min(dist + dash, length)
        draw.line((x1 + ux * dist, y1 + uy * dist, x1 + ux * end, y1 + uy * end), fill=red, width=width)
        dist += dash + gap


def draw_label(draw: ImageDraw.ImageDraw, xy: tuple[int, int], lines: list[str], font: ImageFont.ImageFont) -> None:
    red = (220, 0, 0)
    white = (255, 255, 255)
    black = (0, 0, 0)
    x, y = xy
    padding, gap = 10, 6
    bboxes = [draw.textbbox((0, 0), line, font=font) for line in lines]
    widths = [b[2] - b[0] for b in bboxes]
    heights = [b[3] - b[1] for b in bboxes]
    box_w = max(widths) + padding * 2
    box_h = sum(heights) + gap * (len(lines) - 1) + padding * 2
    draw.rectangle((x, y, x + box_w, y + box_h), fill=white, outline=red, width=2)
    ty = y + padding
    for line, h in zip(lines, heights):
        draw.text((x + padding, ty), line, fill=black, font=font)
        ty += h + gap


def annotate_display(
    source: Path,
    output: Path,
    crops: list[CropSpec],
    fits: list[FitResult],
    cylinder: tuple[int, int, int],
    soil: tuple[int, int, int, int],
    cylinder_mm: float,
    spread_mm: float,
) -> None:
    img = Image.open(source).convert("RGB")
    draw = ImageDraw.Draw(img)
    font = load_font(26)
    cyl_left, cyl_right, cyl_bottom = cylinder
    soil_left, _soil_top, soil_right, soil_bottom = soil

    label_positions = {"left": (80, cyl_bottom + 38), "right": (min(img.width - 520, soil_right - 140), cyl_bottom + 38)}
    for crop, fit in zip(crops, fits):
        x0, x1 = 0, crop.width - 1
        y0 = fit.slope_tan_repose_angle * x0 + fit.intercept_y_up
        y1 = fit.slope_tan_repose_angle * x1 + fit.intercept_y_up
        p0 = (crop.left + x0, crop.top + (crop.height - 1 - y0))
        p1 = (crop.left + x1, crop.top + (crop.height - 1 - y1))
        draw_dashed_line(draw, p0, p1)
        name = "左侧拟合" if fit.side == "left" else "右侧拟合"
        sign = "+" if fit.intercept_y_up >= 0 else "-"
        draw_label(
            draw,
            label_positions[fit.side],
            [
                name,
                f"拟合方程: y = {fit.slope_tan_repose_angle:.6f}x {sign} {abs(fit.intercept_y_up):.6f}",
                f"休止角: {fit.repose_angle_deg:.3f}°",
                f"决定系数: R² = {fit.r_squared:.6f}",
            ],
            font,
        )

    # Cylinder diameter marker just below the cylinder.
    cyl_line_y = cyl_bottom + 28
    draw_dashed_line(draw, (cyl_left, cyl_line_y), (cyl_right, cyl_line_y))
    draw.line((cyl_left, cyl_bottom, cyl_left, cyl_line_y + 10), fill=(220, 0, 0), width=4)
    draw.line((cyl_right, cyl_bottom, cyl_right, cyl_line_y + 10), fill=(220, 0, 0), width=4)
    draw_label(draw, ((cyl_left + cyl_right) // 2 - 120, cyl_line_y + 18), [f"圆筒外径: {cylinder_mm:g} mm"], font)

    # Spread distance marker below soil base.
    spread_line_y = soil_bottom + 22
    draw_dashed_line(draw, (soil_left, spread_line_y), (soil_right, spread_line_y))
    draw.line((soil_left, soil_bottom, soil_left, spread_line_y + 10), fill=(220, 0, 0), width=4)
    draw.line((soil_right, soil_bottom, soil_right, spread_line_y + 10), fill=(220, 0, 0), width=4)
    draw_label(draw, ((soil_left + soil_right) // 2 - 160, spread_line_y + 22), [f"扩散距离: {spread_mm:.3f} mm"], font)

    output.parent.mkdir(exist_ok=True)
    img.save(output)


def group_id_from_stem(stem: str) -> str:
    parts = stem.split("_")
    return parts[0] if parts else stem


def group_sort_key(group_id: str) -> tuple[str, int | str]:
    prefix = group_id.rstrip("0123456789")
    number = group_id[len(prefix) :]
    if number:
        return (prefix, int(number))
    return (group_id, group_id)


def sort_fit_key(row: dict) -> tuple[tuple[str, int | str], str, int]:
    stem = Path(str(row["source_image"])).stem
    side_order = 0 if row.get("side") == "left" else 1
    return (group_sort_key(group_id_from_stem(stem)), stem, side_order)


def sort_spread_key(row: dict) -> tuple[tuple[str, int | str], str]:
    stem = Path(str(row["image"])).stem
    return (group_sort_key(group_id_from_stem(stem)), stem)


def merge_group_column(sheet, start_row: int, end_row: int) -> None:
    row = start_row
    while row <= end_row:
        group = sheet.cell(row=row, column=1).value
        merge_end = row
        while merge_end + 1 <= end_row and sheet.cell(row=merge_end + 1, column=1).value == group:
            merge_end += 1
        if merge_end > row:
            sheet.merge_cells(start_row=row, start_column=1, end_row=merge_end, end_column=1)
        row = merge_end + 1


def write_results(root: Path, fit_rows: list[FitResult], spread_row: dict[str, float | int | str]) -> None:
    json_path = root / "results_data.json"
    existing_fits: list[dict] = []
    if json_path.exists():
        try:
            loaded = json.loads(json_path.read_text(encoding="utf-8"))
            if isinstance(loaded, list):
                existing_fits = [row for row in loaded if isinstance(row, dict) and "source_image" in row]
        except Exception:
            existing_fits = []

    fit_map = {str(row["source_image"]): row for row in existing_fits}
    for row in fit_rows:
        fit_map[row.source_image] = asdict(row)
    all_fits = sorted(fit_map.values(), key=sort_fit_key)
    json_path.write_text(json.dumps(all_fits, ensure_ascii=False, indent=2), encoding="utf-8")

    spread_json = root / "spread_distance_data.json"
    existing_spreads: list[dict] = []
    if spread_json.exists():
        try:
            loaded = json.loads(spread_json.read_text(encoding="utf-8"))
            if isinstance(loaded, list):
                existing_spreads = [row for row in loaded if isinstance(row, dict) and "image" in row]
        except Exception:
            existing_spreads = []
    if not existing_spreads and (root / "results.xlsx").exists():
        try:
            from openpyxl import load_workbook

            prior_wb = load_workbook(root / "results.xlsx", data_only=True)
            if "扩散距离" in prior_wb.sheetnames:
                prior_ws = prior_wb["扩散距离"]
                headers = [cell.value for cell in prior_ws[2]]
                for row in prior_ws.iter_rows(min_row=3, values_only=True):
                    if not row or not row[1]:
                        continue
                    data = dict(zip(headers, row))
                    existing_spreads.append(
                        {
                            "image": data["图片名"],
                            "soil_left": data["土堆左边界(px)"],
                            "soil_right": data["土堆右边界(px)"],
                            "soil_width_px": data["土堆底部长度(px)"],
                            "cyl_left": data["圆筒左边界(px)"],
                            "cyl_right": data["圆筒右边界(px)"],
                            "cyl_width_px": data["圆筒外径(px)"],
                            "cylinder_mm": data["圆筒外径(mm)"],
                            "mm_per_px": data["比例(mm/px)"],
                            "spread_mm": data["扩散距离(mm)"],
                        }
                    )
        except Exception:
            existing_spreads = []
    spread_map = {str(row["image"]): row for row in existing_spreads}
    spread_map[str(spread_row["image"])] = dict(spread_row)
    all_spreads = sorted(spread_map.values(), key=sort_spread_key)
    spread_json.write_text(json.dumps(all_spreads, ensure_ascii=False, indent=2), encoding="utf-8")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception:
        (root / "results.csv").write_text(
            "group_id,source_image,slope_tan_repose_angle,repose_angle_deg,r_squared\n"
            + "\n".join(
                f"{group_id_from_stem(Path(str(r['source_image'])).stem)},{r['source_image']},{r['slope_tan_repose_angle']},{r['repose_angle_deg']},{r['r_squared']}"
                for r in all_fits
            ),
            encoding="utf-8",
        )
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "休止角"
    ws.append(["二值化边界拟合与休止角结果"])
    ws.append(["组号", "截断图片", "二值化图片", "侧别", "宽度(px)", "高度(px)", "阈值", "边界点数量", "斜率 k = tan(休止角)", "休止角(度)", "拟合优度 R²", "截距(y向上)"])
    for row in all_fits:
        group_id = group_id_from_stem(Path(str(row["source_image"])).stem)
        ws.append([
            group_id,
            row["source_image"],
            row["binary_image"],
            row.get("side", ""),
            row["width_px"],
            row["height_px"],
            row["threshold"],
            row["boundary_points"],
            row["slope_tan_repose_angle"],
            row["repose_angle_deg"],
            row["r_squared"],
            row["intercept_y_up"],
        ])
    if all_fits:
        merge_group_column(ws, 3, 2 + len(all_fits))
    ws.append([])
    ws.append(["说明：坐标采用数学坐标系，x 向右、y 向上；斜率 k 为休止角的正切值，休止角按 atan(abs(k)) 计算。"])

    ws2 = wb.create_sheet("扩散距离")
    ws2.append(["黑色土堆底部扩散距离测量"])
    ws2.append(["组号", "图片名", "土堆左边界(px)", "土堆右边界(px)", "土堆底部长度(px)", "圆筒左边界(px)", "圆筒右边界(px)", "圆筒外径(px)", "圆筒外径(mm)", "比例(mm/px)", "扩散距离(mm)"])
    for row in all_spreads:
        group_id = group_id_from_stem(Path(str(row["image"])).stem)
        ws2.append([
            group_id,
            row["image"],
            row["soil_left"],
            row["soil_right"],
            row["soil_width_px"],
            row["cyl_left"],
            row["cyl_right"],
            row["cyl_width_px"],
            row["cylinder_mm"],
            row["mm_per_px"],
            row["spread_mm"],
        ])
    if all_spreads:
        merge_group_column(ws2, 3, 2 + len(all_spreads))
    ws2.append([])
    ws2.append(["说明：扩散距离按黑色土堆最左侧到最右侧的水平距离计算；圆筒外径用于像素到毫米换算。"])

    for sheet in [ws, ws2]:
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=sheet.max_column)
        sheet[1][0].font = Font(bold=True, size=14)
        sheet[1][0].fill = PatternFill("solid", fgColor="D9EAF7")
        for cell in sheet[2]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="EAF3F8")
        for row_cells in sheet.iter_rows():
            for cell in row_cells:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_index in range(1, sheet.max_column + 1):
            letter = sheet.cell(row=2, column=column_index).column_letter
            sheet.column_dimensions[letter].width = 18
        sheet.freeze_panes = "A3"
    wb.save(root / "results.xlsx")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", required=True, help="Project root containing initial_fig.")
    parser.add_argument("--image", default="r1_x.png", help="Image filename under initial_fig.")
    parser.add_argument("--cylinder-mm", type=float, default=95.0)
    parser.add_argument("--threshold", type=int, default=128)
    parser.add_argument("--prefix", default=None)
    args = parser.parse_args()

    root = Path(args.root)
    source = root / "initial_fig" / args.image
    if not source.exists():
        raise FileNotFoundError(source)
    prefix = args.prefix or source.stem
    local_dir = root / "local_fig"
    binary_dir = root / "binary_fig"
    display_dir = root / "display_fig"
    for directory in [local_dir, binary_dir, display_dir]:
        directory.mkdir(exist_ok=True)

    img = Image.open(source).convert("RGB")
    crops, cylinder, soil = make_crops(img, prefix, local_dir)
    fits = [binarize_and_fit(local_dir / crop.filename, binary_dir / crop.filename, args.threshold, crop.side) for crop in crops]

    cyl_left, cyl_right, _ = cylinder
    soil_left, _soil_top, soil_right, _soil_bottom = soil
    cyl_width_px = cyl_right - cyl_left + 1
    soil_width_px = soil_right - soil_left + 1
    mm_per_px = args.cylinder_mm / cyl_width_px
    spread_mm = soil_width_px * mm_per_px
    spread_row = {
        "image": args.image,
        "soil_left": soil_left,
        "soil_right": soil_right,
        "soil_width_px": soil_width_px,
        "cyl_left": cyl_left,
        "cyl_right": cyl_right,
        "cyl_width_px": cyl_width_px,
        "cylinder_mm": args.cylinder_mm,
        "mm_per_px": mm_per_px,
        "spread_mm": spread_mm,
    }
    write_results(root, fits, spread_row)
    annotate_display(source, display_dir / args.image, crops, fits, cylinder, soil, args.cylinder_mm, spread_mm)

    print(json.dumps({"fits": [asdict(row) for row in fits], "spread": spread_row}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
